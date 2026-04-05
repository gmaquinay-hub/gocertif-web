#!/usr/bin/env python3
"""GoCertif Web v2 — Application de gestion des grilles d'audit HAS V2025
Serveur : Tornado (stdlib-compatible)
v2 : authentification, historique par établissement, plan d'actions, import/export JSON
"""

import json
import os
import sqlite3
import io
import hashlib
import hmac
import secrets
import base64
from datetime import datetime
from urllib.parse import urlparse, parse_qs

import tornado.web
import tornado.ioloop
import tornado.httpserver

# ─── Configuration ────────────────────────────────────────────────────────────
BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
DB_PATH       = os.path.join(BASE_DIR, 'gocertif.db')
PORT          = int(os.environ.get('PORT', 5050))
COOKIE_SECRET = os.environ.get('COOKIE_SECRET', secrets.token_hex(32))

# ─── Chargement des critères V2025 ────────────────────────────────────────────
with open(os.path.join(BASE_DIR, 'criteria_v2025.json'), encoding='utf-8') as f:
    CRITERIA = json.load(f)

CRITERIA_BY_NUM = {c['numero']: c for c in CRITERIA}

# ─── Base de données ───────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            email             TEXT NOT NULL UNIQUE,
            password_hash     TEXT NOT NULL,
            nom_etablissement TEXT DEFAULT '',
            type_structure    TEXT DEFAULT '',
            created_at        TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS evaluations (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id       INTEGER REFERENCES users(id) ON DELETE CASCADE,
            nom           TEXT NOT NULL DEFAULT 'Évaluation',
            methode       TEXT NOT NULL DEFAULT 'Parcours traceur',
            etablissement TEXT DEFAULT '',
            service       TEXT DEFAULT '',
            date_audit    TEXT DEFAULT '',
            auditeurs     TEXT DEFAULT '',
            code_calista  TEXT DEFAULT '',
            secteur_pec   TEXT DEFAULT '',
            mode_pec      TEXT DEFAULT '',
            mode_entree   TEXT DEFAULT '',
            age_patient   TEXT DEFAULT '',
            caracteristiques TEXT DEFAULT '',
            type_parcours TEXT DEFAULT '',
            notes_globales   TEXT DEFAULT '',
            created_at    TEXT DEFAULT (datetime('now')),
            updated_at    TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS reponses (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            evaluation_id  INTEGER NOT NULL,
            critere_numero TEXT NOT NULL,
            reponse        TEXT CHECK(reponse IN ('Oui','Non','NA','')),
            notes          TEXT DEFAULT '',
            plan_action    TEXT DEFAULT '',
            updated_at     TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (evaluation_id) REFERENCES evaluations(id) ON DELETE CASCADE,
            UNIQUE(evaluation_id, critere_numero)
        );

        CREATE TABLE IF NOT EXISTS reponses_ee (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            evaluation_id  INTEGER NOT NULL,
            critere_numero TEXT NOT NULL,
            ee_index       INTEGER NOT NULL,
            reponse        TEXT DEFAULT '' CHECK(reponse IN ('Oui','Non','NA','')),
            updated_at     TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (evaluation_id) REFERENCES evaluations(id) ON DELETE CASCADE,
            UNIQUE(evaluation_id, critere_numero, ee_index)
        );

        CREATE TABLE IF NOT EXISTS actions (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id       INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            evaluation_id INTEGER REFERENCES evaluations(id) ON DELETE SET NULL,
            critere_num   TEXT DEFAULT '',
            titre         TEXT NOT NULL DEFAULT 'Action sans titre',
            description   TEXT DEFAULT '',
            responsable   TEXT DEFAULT '',
            echeance      TEXT DEFAULT '',
            statut        TEXT DEFAULT 'todo' CHECK(statut IN ('todo','in_progress','done')),
            priorite      TEXT DEFAULT 'medium' CHECK(priorite IN ('low','medium','high','critical')),
            created_at    TEXT DEFAULT (datetime('now')),
            updated_at    TEXT DEFAULT (datetime('now'))
        );
    ''')
    conn.commit()
    conn.close()

init_db()

# ─── Seed admin ───────────────────────────────────────────────────────────────
def seed_admin():
    """Crée le compte admin au démarrage si ADMIN_EMAIL / ADMIN_PASSWORD sont définis
    et que l'email n'existe pas encore en base."""
    email = os.environ.get('ADMIN_EMAIL', '').strip().lower()
    pwd   = os.environ.get('ADMIN_PASSWORD', '').strip()
    nom   = os.environ.get('ADMIN_NOM', 'Établissement')
    if not email or not pwd:
        return
    conn = get_db()
    existing = conn.execute('SELECT id FROM users WHERE email=?', (email,)).fetchone()
    if not existing:
        ph = hash_password(pwd)
        conn.execute(
            'INSERT INTO users (email, password_hash, nom_etablissement, type_structure) VALUES (?,?,?,?)',
            (email, ph, nom, 'Administration')
        )
        conn.commit()
        print(f'  ✔  Compte admin créé : {email}')
    conn.close()

# ─── Auth helpers ──────────────────────────────────────────────────────────────
def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    dk   = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 200_000)
    return f"{salt}:{base64.b64encode(dk).decode()}"

def verify_password(password: str, stored: str) -> bool:
    try:
        salt, encoded = stored.split(':')
        dk = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 200_000)
        return hmac.compare_digest(base64.b64encode(dk).decode(), encoded)
    except Exception:
        return False

# ─── Scoring ──────────────────────────────────────────────────────────────────
def calculate_score(evaluation_id):
    conn = get_db()
    rows = conn.execute(
        'SELECT critere_numero, reponse FROM reponses WHERE evaluation_id = ?',
        (evaluation_id,)
    ).fetchall()
    ee_rows = conn.execute(
        'SELECT critere_numero, ee_index, reponse FROM reponses_ee WHERE evaluation_id = ?',
        (evaluation_id,)
    ).fetchall()
    conn.close()

    rep = {r['critere_numero']: r['reponse'] for r in rows}

    # Agrège les réponses EE par critère → déduit la réponse critère
    ee_by_crit = {}
    for r in ee_rows:
        ee_by_crit.setdefault(r['critere_numero'], []).append(r['reponse'])

    def aggregate_ee(ee_list):
        oui_c = ee_list.count('Oui')
        non_c = ee_list.count('Non')
        na_c  = ee_list.count('NA')
        evaluated = oui_c + non_c
        if evaluated == 0:
            return 'NA' if na_c == len(ee_list) else ''
        return 'Oui' if non_c == 0 else 'Non'

    oui = non = na = nr = 0
    fiches_anomalie = []

    for c in CRITERIA:
        if c['numero'] in ee_by_crit:
            r = aggregate_ee(ee_by_crit[c['numero']])
        else:
            r = rep.get(c['numero'], '')
        if r == 'Oui':   oui += 1
        elif r == 'Non':
            non += 1
            if c['niveau'] == 'Impératif':
                fiches_anomalie.append(c['numero'])
        elif r == 'NA':  na  += 1
        else:            nr  += 1

    total = oui + non
    pct   = round(oui / total * 100, 1) if total else None

    if pct is None:
        label = 'Non évalué'; cls = 'gray'
    elif pct >= 80 and len(fiches_anomalie) == 0:
        label = 'Conforme — Certifié'; cls = 'green'
    elif pct >= 60 and len(fiches_anomalie) <= 2:
        label = 'Partiellement conforme'; cls = 'orange'
    else:
        label = 'Non conforme — Risque de non-certification'; cls = 'red'

    return {
        'oui': oui, 'non': non, 'na': na, 'non_renseigne': nr,
        'total_evalue': total, 'score_pct': pct,
        'nb_anomalies': len(fiches_anomalie), 'fiches_anomalie': fiches_anomalie,
        'niveau_decision': label, 'decision_class': cls,
        'total_criteres': len(CRITERIA)
    }

# ─── Base Handler ──────────────────────────────────────────────────────────────
class BaseHandler(tornado.web.RequestHandler):
    def set_default_headers(self):
        self.set_header('Access-Control-Allow-Origin', '*')
        self.set_header('Access-Control-Allow-Headers', 'Content-Type')
        self.set_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')

    def options(self, *args):
        self.set_status(204)
        self.finish()

    def json(self, data, status=200):
        self.set_status(status)
        self.set_header('Content-Type', 'application/json; charset=utf-8')
        self.write(json.dumps(data, ensure_ascii=False))

    def body_json(self):
        try:
            return json.loads(self.request.body)
        except Exception:
            return {}

    def get_current_user(self):
        user_id = self.get_secure_cookie('user_id')
        if not user_id:
            return None
        conn = get_db()
        user = conn.execute('SELECT id, email, nom_etablissement, type_structure FROM users WHERE id=?',
                            (int(user_id),)).fetchone()
        conn.close()
        return dict(user) if user else None

    def require_auth(self):
        """Retourne l'utilisateur ou envoie 401. Utiliser avec `user = self.require_auth(); if not user: return`"""
        user = self.get_current_user()
        if not user:
            self.json({'error': 'Authentification requise', 'code': 'UNAUTHORIZED'}, 401)
        return user

# ─── Auth Handlers ─────────────────────────────────────────────────────────────
class AuthRegisterHandler(BaseHandler):
    def post(self):
        data  = self.body_json()
        email = (data.get('email') or '').strip().lower()
        pwd   = data.get('password', '')
        nom   = (data.get('nom_etablissement') or '').strip()
        typ   = (data.get('type_structure') or '').strip()

        if not email or not pwd:
            return self.json({'error': 'Email et mot de passe requis'}, 400)
        if len(pwd) < 6:
            return self.json({'error': 'Mot de passe trop court (6 caractères minimum)'}, 400)

        conn = get_db()
        existing = conn.execute('SELECT id FROM users WHERE email=?', (email,)).fetchone()
        if existing:
            conn.close()
            return self.json({'error': 'Cet email est déjà utilisé'}, 409)

        ph = hash_password(pwd)
        cur = conn.execute(
            'INSERT INTO users (email, password_hash, nom_etablissement, type_structure) VALUES (?,?,?,?)',
            (email, ph, nom, typ)
        )
        user_id = cur.lastrowid
        conn.commit()
        conn.close()

        self.set_secure_cookie('user_id', str(user_id), expires_days=30, samesite='Lax')
        self.json({
            'id': user_id, 'email': email,
            'nom_etablissement': nom, 'type_structure': typ,
            'message': 'Compte créé avec succès'
        }, 201)

class AuthLoginHandler(BaseHandler):
    def post(self):
        data  = self.body_json()
        email = (data.get('email') or '').strip().lower()
        pwd   = data.get('password', '')

        conn  = get_db()
        user  = conn.execute('SELECT * FROM users WHERE email=?', (email,)).fetchone()
        conn.close()

        if not user or not verify_password(pwd, user['password_hash']):
            return self.json({'error': 'Email ou mot de passe incorrect'}, 401)

        self.set_secure_cookie('user_id', str(user['id']), expires_days=30, samesite='Lax')
        self.json({
            'id': user['id'], 'email': user['email'],
            'nom_etablissement': user['nom_etablissement'],
            'type_structure': user['type_structure']
        })

class AuthLogoutHandler(BaseHandler):
    def post(self):
        self.clear_cookie('user_id')
        self.json({'message': 'Déconnecté'})

class AuthMeHandler(BaseHandler):
    def get(self):
        user = self.get_current_user()
        if not user:
            return self.json({'authenticated': False}, 200)
        self.json({'authenticated': True, **user})

    def put(self):
        user = self.require_auth()
        if not user: return
        data = self.body_json()
        nom  = (data.get('nom_etablissement') or '').strip()
        typ  = (data.get('type_structure') or '').strip()
        conn = get_db()
        conn.execute('UPDATE users SET nom_etablissement=?, type_structure=? WHERE id=?',
                     (nom, typ, user['id']))
        conn.commit()
        conn.close()
        self.json({'message': 'Profil mis à jour'})

# ─── Index ─────────────────────────────────────────────────────────────────────
class IndexHandler(BaseHandler):
    def get(self):
        with open(os.path.join(BASE_DIR, 'templates', 'index.html'), encoding='utf-8') as f:
            self.set_header('Content-Type', 'text/html; charset=utf-8')
            self.write(f.read())

# ─── Criteria ──────────────────────────────────────────────────────────────────
class CriteriaHandler(BaseHandler):
    def get(self):
        methode = self.get_argument('methode', '')
        champ   = self.get_argument('champ', '')
        niveau  = self.get_argument('niveau', '')
        result  = CRITERIA
        if methode and methode != 'Toutes':
            result = [c for c in result if methode in c.get('methodes', [])]
        if champ and champ != 'Tous':
            result = [c for c in result if c['champ'] == champ or c['champ'] == "Tout l'établissement"]
        if niveau and niveau != 'Tous':
            result = [c for c in result if c['niveau'] == niveau]
        self.json(result)

class CriteriaMetaHandler(BaseHandler):
    def get(self):
        self.json({
            'chapitres': list(dict.fromkeys(c['chapitre'] for c in CRITERIA)),
            'objectifs': list(dict.fromkeys(f"{c['objectif_num']} – {c['objectif']}" for c in CRITERIA)),
            'champs':    sorted(set(c['champ'] for c in CRITERIA)),
            'methodes':  sorted(set(m for c in CRITERIA for m in c.get('methodes', []))),
            'niveaux':   ['Impératif', 'Standard', 'Avancé'],
            'secteurs_pec': ['Chirurgie et interventionnel','Maternité','Médecine','Psychiatrie et santé mentale',
                             'Soins critiques','Soins de suite et réadaptation','Soins de longue durée'],
            'modes_pec': ['Ambulatoire','Hospitalisation complète','HAD','HDJ','HDS'],
            'modes_entree': ['Programmé','SAMU-SMUR','Urgences'],
            'ages_patient': ['Adulte','Enfant et adolescent','Patient âgé (+ 75 ans)'],
            'caracteristiques': ['Maladie chronique','Pas de situation particulière',"Patient atteint d'un cancer",
                                  'Patient en situation de handicap','Patient en situation de précarité'],
        })

# ─── Evaluations ───────────────────────────────────────────────────────────────
class EvaluationsHandler(BaseHandler):
    def get(self):
        user = self.require_auth()
        if not user: return
        conn  = get_db()
        evals = conn.execute(
            'SELECT * FROM evaluations WHERE user_id=? ORDER BY updated_at DESC',
            (user['id'],)
        ).fetchall()
        conn.close()
        result = []
        for e in evals:
            d = dict(e)
            d['score'] = calculate_score(e['id'])
            result.append(d)
        self.json(result)

    def post(self):
        user = self.require_auth()
        if not user: return
        data = self.body_json()
        conn = get_db()
        cur  = conn.execute('''
            INSERT INTO evaluations
            (user_id, nom, methode, etablissement, service, date_audit, auditeurs, code_calista,
             secteur_pec, mode_pec, mode_entree, age_patient, caracteristiques, type_parcours, notes_globales)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            user['id'],
            data.get('nom','Nouvelle évaluation'), data.get('methode','Parcours traceur'),
            data.get('etablissement',''), data.get('service',''), data.get('date_audit',''),
            data.get('auditeurs',''), data.get('code_calista',''), data.get('secteur_pec',''),
            data.get('mode_pec',''), data.get('mode_entree',''), data.get('age_patient',''),
            data.get('caracteristiques',''), data.get('type_parcours',''), data.get('notes_globales','')
        ))
        eval_id = cur.lastrowid
        conn.commit()
        conn.close()
        self.json({'id': eval_id, 'message': 'Évaluation créée'}, 201)

class EvaluationHandler(BaseHandler):
    def _check_owner(self, eval_id, user_id):
        conn = get_db()
        e = conn.execute('SELECT * FROM evaluations WHERE id=? AND user_id=?', (eval_id, user_id)).fetchone()
        conn.close()
        return e

    def get(self, eval_id):
        user = self.require_auth()
        if not user: return
        conn = get_db()
        e = conn.execute('SELECT * FROM evaluations WHERE id=? AND user_id=?',
                         (eval_id, user['id'])).fetchone()
        if not e:
            conn.close()
            return self.json({'error': 'Non trouvé'}, 404)
        reps = conn.execute(
            'SELECT critere_numero, reponse, notes, plan_action FROM reponses WHERE evaluation_id=?',
            (eval_id,)
        ).fetchall()
        conn.close()
        self.json({**dict(e), 'reponses': {r['critere_numero']: dict(r) for r in reps},
                   'score': calculate_score(int(eval_id))})

    def put(self, eval_id):
        user = self.require_auth()
        if not user: return
        data = self.body_json()
        conn = get_db()
        conn.execute('''
            UPDATE evaluations SET nom=?,methode=?,etablissement=?,service=?,date_audit=?,
            auditeurs=?,code_calista=?,secteur_pec=?,mode_pec=?,mode_entree=?,age_patient=?,
            caracteristiques=?,type_parcours=?,notes_globales=?,updated_at=datetime('now')
            WHERE id=? AND user_id=?
        ''', (data.get('nom'), data.get('methode'), data.get('etablissement'), data.get('service'),
              data.get('date_audit'), data.get('auditeurs'), data.get('code_calista'),
              data.get('secteur_pec'), data.get('mode_pec'), data.get('mode_entree'),
              data.get('age_patient'), data.get('caracteristiques'), data.get('type_parcours'),
              data.get('notes_globales'), eval_id, user['id']))
        conn.commit()
        conn.close()
        self.json({'message': 'Mis à jour'})

    def delete(self, eval_id):
        user = self.require_auth()
        if not user: return
        conn = get_db()
        conn.execute('DELETE FROM reponses WHERE evaluation_id=?', (eval_id,))
        conn.execute('DELETE FROM evaluations WHERE id=? AND user_id=?', (eval_id, user['id']))
        conn.commit()
        conn.close()
        self.json({'message': 'Supprimé'})

# ─── Réponses ──────────────────────────────────────────────────────────────────
class ReponsesHandler(BaseHandler):
    def post(self, eval_id):
        user = self.require_auth()
        if not user: return
        data = self.body_json()
        conn = get_db()
        conn.execute('''
            INSERT INTO reponses (evaluation_id, critere_numero, reponse, notes, plan_action, updated_at)
            VALUES (?,?,?,?,?,datetime('now'))
            ON CONFLICT(evaluation_id, critere_numero) DO UPDATE SET
                reponse=excluded.reponse, notes=excluded.notes,
                plan_action=excluded.plan_action, updated_at=datetime('now')
        ''', (eval_id, data.get('critere_numero'), data.get('reponse',''),
               data.get('notes',''), data.get('plan_action','')))
        conn.execute("UPDATE evaluations SET updated_at=datetime('now') WHERE id=?", (eval_id,))
        conn.commit()
        conn.close()
        self.json({'message': 'Enregistré', 'score': calculate_score(int(eval_id))})

class BulkReponsesHandler(BaseHandler):
    def post(self, eval_id):
        user = self.require_auth()
        if not user: return
        data     = self.body_json()
        reponses = data.get('reponses', [])
        conn     = get_db()
        for rep in reponses:
            conn.execute('''
                INSERT INTO reponses (evaluation_id, critere_numero, reponse, notes, plan_action, updated_at)
                VALUES (?,?,?,?,?,datetime('now'))
                ON CONFLICT(evaluation_id, critere_numero) DO UPDATE SET
                    reponse=excluded.reponse, notes=excluded.notes,
                    plan_action=excluded.plan_action, updated_at=datetime('now')
            ''', (eval_id, rep.get('critere_numero'), rep.get('reponse',''),
                   rep.get('notes',''), rep.get('plan_action','')))
        conn.execute("UPDATE evaluations SET updated_at=datetime('now') WHERE id=?", (eval_id,))
        conn.commit()
        conn.close()
        self.json({'message': f'{len(reponses)} réponses enregistrées', 'score': calculate_score(int(eval_id))})

class ScoreHandler(BaseHandler):
    def get(self, eval_id):
        self.json(calculate_score(int(eval_id)))

# ─── Import / Export JSON ──────────────────────────────────────────────────────
class ExportJSONHandler(BaseHandler):
    def get(self, eval_id):
        user = self.require_auth()
        if not user: return
        conn = get_db()
        e    = conn.execute('SELECT * FROM evaluations WHERE id=? AND user_id=?',
                            (eval_id, user['id'])).fetchone()
        if not e:
            conn.close()
            return self.json({'error': 'Non trouvé'}, 404)
        reps = conn.execute(
            'SELECT critere_numero, reponse, notes, plan_action FROM reponses WHERE evaluation_id=?',
            (eval_id,)
        ).fetchall()
        conn.close()

        export = {
            'version': '2',
            'app': 'GoCertif Web',
            'referentiel': 'HAS V2025',
            'exported_at': datetime.now().isoformat(),
            'evaluation': dict(e),
            'reponses': [dict(r) for r in reps],
            'score': calculate_score(int(eval_id))
        }
        nom      = (e['nom'] or 'eval').replace(' ','_')
        filename = f"GoCertif_{nom}_{datetime.now():%Y%m%d}.json"
        self.set_header('Content-Type', 'application/json; charset=utf-8')
        self.set_header('Content-Disposition', f'attachment; filename="{filename}"')
        self.write(json.dumps(export, ensure_ascii=False, indent=2))

class ImportJSONHandler(BaseHandler):
    def post(self):
        user = self.require_auth()
        if not user: return
        data = self.body_json()

        if data.get('app') != 'GoCertif Web':
            return self.json({'error': 'Format de fichier non reconnu'}, 400)

        ev = data.get('evaluation', {})
        reponses = data.get('reponses', [])

        conn = get_db()
        cur  = conn.execute('''
            INSERT INTO evaluations
            (user_id, nom, methode, etablissement, service, date_audit, auditeurs, code_calista,
             secteur_pec, mode_pec, mode_entree, age_patient, caracteristiques, type_parcours, notes_globales)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            user['id'],
            ev.get('nom','Import'), ev.get('methode','Parcours traceur'),
            ev.get('etablissement',''), ev.get('service',''), ev.get('date_audit',''),
            ev.get('auditeurs',''), ev.get('code_calista',''), ev.get('secteur_pec',''),
            ev.get('mode_pec',''), ev.get('mode_entree',''), ev.get('age_patient',''),
            ev.get('caracteristiques',''), ev.get('type_parcours',''), ev.get('notes_globales','')
        ))
        new_id = cur.lastrowid

        for rep in reponses:
            if rep.get('critere_numero') in CRITERIA_BY_NUM:
                conn.execute('''
                    INSERT OR REPLACE INTO reponses
                    (evaluation_id, critere_numero, reponse, notes, plan_action, updated_at)
                    VALUES (?,?,?,?,?,datetime('now'))
                ''', (new_id, rep['critere_numero'], rep.get('reponse',''),
                      rep.get('notes',''), rep.get('plan_action','')))
        conn.commit()
        conn.close()
        self.json({'id': new_id, 'message': f'Évaluation importée ({len(reponses)} réponses)'}, 201)

# ─── Plan d'actions ────────────────────────────────────────────────────────────
class ActionsHandler(BaseHandler):
    def get(self):
        user = self.require_auth()
        if not user: return
        eval_id = self.get_argument('evaluation_id', None)
        statut  = self.get_argument('statut', None)
        conn    = get_db()
        query   = 'SELECT * FROM actions WHERE user_id=?'
        params  = [user['id']]
        if eval_id:
            query  += ' AND evaluation_id=?'; params.append(eval_id)
        if statut:
            query  += ' AND statut=?'; params.append(statut)
        query += ' ORDER BY priorite DESC, echeance ASC, created_at DESC'
        rows  = conn.execute(query, params).fetchall()
        conn.close()
        self.json([dict(r) for r in rows])

    def post(self):
        user = self.require_auth()
        if not user: return
        data = self.body_json()
        conn = get_db()
        cur  = conn.execute('''
            INSERT INTO actions
            (user_id, evaluation_id, critere_num, titre, description, responsable,
             echeance, statut, priorite)
            VALUES (?,?,?,?,?,?,?,?,?)
        ''', (
            user['id'],
            data.get('evaluation_id'), data.get('critere_num',''),
            data.get('titre','Action'), data.get('description',''),
            data.get('responsable',''), data.get('echeance',''),
            data.get('statut','todo'), data.get('priorite','medium')
        ))
        action_id = cur.lastrowid
        conn.commit()
        conn.close()
        self.json({'id': action_id, 'message': 'Action créée'}, 201)

class ActionHandler(BaseHandler):
    def put(self, action_id):
        user = self.require_auth()
        if not user: return
        data = self.body_json()
        conn = get_db()
        conn.execute('''
            UPDATE actions SET titre=?, description=?, responsable=?, echeance=?,
            statut=?, priorite=?, critere_num=?, updated_at=datetime('now')
            WHERE id=? AND user_id=?
        ''', (data.get('titre'), data.get('description'), data.get('responsable'),
              data.get('echeance'), data.get('statut'), data.get('priorite'),
              data.get('critere_num',''), action_id, user['id']))
        conn.commit()
        conn.close()
        self.json({'message': 'Action mise à jour'})

    def delete(self, action_id):
        user = self.require_auth()
        if not user: return
        conn = get_db()
        conn.execute('DELETE FROM actions WHERE id=? AND user_id=?', (action_id, user['id']))
        conn.commit()
        conn.close()
        self.json({'message': 'Action supprimée'})

# ─── Score evolution ───────────────────────────────────────────────────────────
class ScoreHistoryHandler(BaseHandler):
    """Retourne l'évolution du score pour toutes les évaluations de l'utilisateur."""
    def get(self):
        user = self.require_auth()
        if not user: return
        conn  = get_db()
        evals = conn.execute(
            'SELECT id, nom, date_audit, created_at FROM evaluations WHERE user_id=? ORDER BY COALESCE(NULLIF(date_audit,""), created_at) ASC',
            (user['id'],)
        ).fetchall()
        conn.close()
        history = []
        for e in evals:
            s = calculate_score(e['id'])
            history.append({
                'id': e['id'], 'nom': e['nom'],
                'date': e['date_audit'] or e['created_at'][:10],
                'score_pct': s['score_pct'],
                'niveau_decision': s['niveau_decision'],
                'decision_class': s['decision_class'],
                'nb_anomalies': s['nb_anomalies'],
                'oui': s['oui'], 'non': s['non'], 'na': s['na']
            })
        self.json(history)

# ─── Export Excel ──────────────────────────────────────────────────────────────
class ExportExcelHandler(BaseHandler):
    def get(self, eval_id):
        user = self.require_auth()
        if not user: return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return self.json({'error': 'openpyxl non installé'}, 500)

        conn = get_db()
        e    = conn.execute('SELECT * FROM evaluations WHERE id=? AND user_id=?',
                            (eval_id, user['id'])).fetchone()
        reps = conn.execute('SELECT critere_numero, reponse, notes, plan_action FROM reponses WHERE evaluation_id=?', (eval_id,)).fetchall()
        conn.close()
        if not e: return self.json({'error': 'Non trouvé'}, 404)

        rep_dict = {r['critere_numero']: dict(r) for r in reps}
        score    = calculate_score(int(eval_id))

        wb = Workbook()
        ws = wb.active
        ws.title = "Grille d'audit"

        def hdr(cell, bg='1E4FA5', fg='FFFFFF'):
            cell.font      = Font(bold=True, color=fg, size=10)
            cell.fill      = PatternFill('solid', fgColor=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        def brd():
            s = Side(style='thin', color='CCCCCC')
            return Border(left=s, right=s, top=s, bottom=s)

        ws.merge_cells('A1:I1')
        c = ws['A1']
        c.value = "GRILLE D'AUDIT HAS V2025 — CERTIFICATION DES ÉTABLISSEMENTS DE SANTÉ"
        c.font = Font(bold=True, color='FFFFFF', size=13)
        c.fill = PatternFill('solid', fgColor='1E4FA5')
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        infos = [
            ('Méthode', e['methode']), ('Établissement', e['etablissement'] or ''),
            ('Service', e['service'] or ''), ("Date de l'audit", e['date_audit'] or ''),
            ('Auditeurs', e['auditeurs'] or ''), ('Code Calista', e['code_calista'] or ''),
        ]
        for extra in [('Secteur PEC', 'secteur_pec'), ('Mode PEC', 'mode_pec'),
                      ('Mode entrée', 'mode_entree'), ('Âge patient', 'age_patient')]:
            if e[extra[1]]:
                infos.append((extra[0], e[extra[1]]))

        r = 2
        for lbl, val in infos:
            ws.merge_cells(f'A{r}:C{r}'); ws.merge_cells(f'D{r}:I{r}')
            ws[f'A{r}'].value = lbl
            ws[f'A{r}'].font  = Font(bold=True, size=10)
            ws[f'A{r}'].fill  = PatternFill('solid', fgColor='E8F0FD')
            ws[f'D{r}'].value = val
            ws[f'D{r}'].font  = Font(size=10)
            ws.row_dimensions[r].height = 17
            r += 1

        r += 1
        ws.merge_cells(f'A{r}:C{r}'); ws.merge_cells(f'D{r}:F{r}'); ws.merge_cells(f'G{r}:I{r}')
        ws[f'A{r}'].value = 'SCORE GLOBAL'
        ws[f'A{r}'].font  = Font(bold=True, size=11)
        ws[f'A{r}'].fill  = PatternFill('solid', fgColor='FFE066')
        spct = f"{score['score_pct']}%" if score['score_pct'] is not None else 'Non évalué'
        ws[f'D{r}'].value = f"Score : {spct}  |  Oui : {score['oui']}  |  Non : {score['non']}  |  NA : {score['na']}"
        ws[f'D{r}'].font  = Font(bold=True, size=11)
        dc = {'green':'00B050','orange':'FF8C00','red':'FF0000','gray':'808080'}.get(score['decision_class'],'808080')
        ws[f'G{r}'].value = f"Fiches anomalie : {score['nb_anomalies']}  |  {score['niveau_decision']}"
        ws[f'G{r}'].font  = Font(bold=True, size=11, color=dc)
        ws.row_dimensions[r].height = 22
        r += 2

        hdrs = ['Chapitre','Objectif','N° Critère','Critère','Champ d\'application','Niveau','Réponse','Notes / Précisions','Plan d\'actions']
        widths = [18, 28, 10, 55, 22, 10, 10, 35, 35]
        for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
            hdr(ws.cell(r, ci, h))
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[r].height = 22
        r += 1

        chap_fill = {'1':'EBF3FA','2':'EAF5EA','3':'FFF5E6'}
        for crit in CRITERIA:
            rd   = rep_dict.get(crit['numero'], {})
            rep  = rd.get('reponse','') or ''
            note = rd.get('notes','')   or ''
            plan = rd.get('plan_action','') or ''
            bg   = chap_fill.get(crit['chapitre_num'],'FFFFFF')

            vals = [crit['chapitre'], crit['objectif'], crit['numero'], crit['titre'],
                    crit['champ'], crit['niveau'], rep, note, plan]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(r, ci, v)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = brd()
                cell.font   = Font(size=9)
                if ci <= 6:
                    cell.fill = PatternFill('solid', fgColor=bg)
                if ci == 6:
                    if v == 'Impératif':
                        cell.font = Font(size=9, bold=True, color='FF0000')
                    elif v == 'Avancé':
                        cell.font = Font(size=9, bold=True, color='7030A0')
                if ci == 7:
                    colors_ = {'Oui':'00B050','Non':'FF0000','NA':'808080'}
                    cell.font = Font(size=9, bold=True, color=colors_.get(v,'000000'))
                    if v == 'Non' and crit['niveau'] == 'Impératif':
                        cell.fill = PatternFill('solid', fgColor='FFE0E0')
            ws.row_dimensions[r].height = 40
            r += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        nom      = (e['nom'] or 'eval').replace(' ','_')
        filename = f"GoCertif_{nom}_{datetime.now():%Y%m%d}.xlsx"
        self.set_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.set_header('Content-Disposition', f'attachment; filename="{filename}"')
        self.write(output.read())

# ─── Export PDF ────────────────────────────────────────────────────────────────
class ExportPDFHandler(BaseHandler):
    def get(self, eval_id):
        user = self.require_auth()
        if not user: return
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm, mm
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Table,
                                             TableStyle, Spacer, HRFlowable)
            from reportlab.lib.enums import TA_CENTER
        except ImportError:
            return self.json({'error': 'reportlab non installé'}, 500)

        conn = get_db()
        e    = conn.execute('SELECT * FROM evaluations WHERE id=? AND user_id=?',
                            (eval_id, user['id'])).fetchone()
        reps = conn.execute('SELECT critere_numero,reponse,notes,plan_action FROM reponses WHERE evaluation_id=?', (eval_id,)).fetchall()
        # Fetch actions for this evaluation
        acts = conn.execute(
            'SELECT * FROM actions WHERE evaluation_id=? ORDER BY priorite DESC, echeance ASC',
            (eval_id,)
        ).fetchall()
        conn.close()
        if not e: return self.json({'error': 'Non trouvé'}, 404)

        rep_dict = {r['critere_numero']: dict(r) for r in reps}
        score    = calculate_score(int(eval_id))

        output = io.BytesIO()
        doc    = SimpleDocTemplate(output, pagesize=landscape(A4),
                                   leftMargin=1*cm, rightMargin=1*cm,
                                   topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        small  = ParagraphStyle('s', parent=styles['Normal'], fontSize=7, leading=9)
        sb     = ParagraphStyle('sb', parent=small, fontName='Helvetica-Bold')
        tiny   = ParagraphStyle('t', parent=styles['Normal'], fontSize=6, leading=8)
        h2     = ParagraphStyle('h2', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold',
                                 spaceBefore=6, spaceAfter=3)

        BLUE  = colors.HexColor('#1E4FA5')
        LBLUE = colors.HexColor('#E8F0FD')

        story = []
        story.append(Paragraph("<font color='#1E4FA5' size='13'><b>GRILLE D'AUDIT HAS V2025</b></font>",
                                ParagraphStyle('tit', alignment=TA_CENTER, spaceAfter=2)))
        story.append(HRFlowable(width='100%', thickness=2, color=BLUE))
        story.append(Spacer(1, 3*mm))

        dc    = {'green':'#00B050','orange':'#FF8C00','red':'#FF0000','gray':'#808080'}.get(score['decision_class'],'#808080')
        spct  = f"{score['score_pct']}%" if score['score_pct'] is not None else 'Non évalué'
        idata = [[
            Paragraph(f"<b>Méthode:</b> {e['methode']}", small),
            Paragraph(f"<b>Établissement:</b> {e['etablissement'] or ''}", small),
            Paragraph(f"<b>Service:</b> {e['service'] or ''}", small),
            Paragraph(f"<b>Date:</b> {e['date_audit'] or ''}", small),
        ],[
            Paragraph(f"<b>Auditeurs:</b> {e['auditeurs'] or ''}", small),
            Paragraph(f"<b>Code Calista:</b> {e['code_calista'] or ''}", small),
            Paragraph(f"<b>Score: <font color='{dc}'>{spct}</font></b> | Oui:{score['oui']} Non:{score['non']} NA:{score['na']}", small),
            Paragraph(f"<b>Fiches anomalie: {score['nb_anomalies']}</b> | <font color='{dc}'>{score['niveau_decision']}</font>", small),
        ]]
        it = Table(idata, colWidths=[6.5*cm]*4)
        it.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,-1),LBLUE),
            ('BOX',(0,0),(-1,-1),0.5,BLUE),
            ('INNERGRID',(0,0),(-1,-1),0.3,colors.lightblue),
            ('VALIGN',(0,0),(-1,-1),'TOP'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
        ]))
        story.append(it)
        story.append(Spacer(1, 3*mm))

        # Grille principale
        cw  = [3.5*cm, 4.5*cm, 1.8*cm, 8*cm, 2.5*cm, 1.8*cm, 1.5*cm, 3.5*cm, 3.5*cm]
        hdrs_row = [Paragraph(f'<b>{h}</b>', sb) for h in
                ['Chapitre','Objectif','N°','Critère','Champ','Niveau','Réponse','Notes','Plan d\'actions']]

        tdata = [hdrs_row]
        tstyles = [
            ('BACKGROUND',(0,0),(-1,0),BLUE),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTSIZE',(0,0),(-1,-1),6),
            ('VALIGN',(0,0),(-1,-1),'TOP'),
            ('INNERGRID',(0,0),(-1,-1),0.3,colors.lightgrey),
            ('BOX',(0,0),(-1,-1),0.5,colors.grey),
        ]

        for i, crit in enumerate(CRITERIA, 1):
            rd   = rep_dict.get(crit['numero'], {})
            rep  = (rd.get('reponse','') or '').strip()
            note = (rd.get('notes','') or '')[:80]
            plan = (rd.get('plan_action','') or '')[:80]

            nc = '#FF0000' if crit['niveau']=='Impératif' else ('#7030A0' if crit['niveau']=='Avancé' else '#333333')
            rc = '#00B050' if rep=='Oui' else ('#FF0000' if rep=='Non' else '#808080')

            tdata.append([
                Paragraph(crit['chapitre'][:20], tiny),
                Paragraph(crit['objectif'][:50], tiny),
                Paragraph(f'<b>{crit["numero"]}</b>', tiny),
                Paragraph(crit['titre'][:120], tiny),
                Paragraph(crit['champ'][:28], tiny),
                Paragraph(f'<font color="{nc}"><b>{crit["niveau"]}</b></font>', tiny),
                Paragraph(f'<font color="{rc}"><b>{rep}</b></font>', tiny),
                Paragraph(note, tiny),
                Paragraph(plan, tiny),
            ])
            if rep == 'Non' and crit['niveau'] == 'Impératif':
                tstyles.append(('BACKGROUND',(6,i),(6,i),colors.HexColor('#FFE0E0')))

        ct = Table(tdata, colWidths=cw, repeatRows=1)
        ct.setStyle(TableStyle(tstyles))
        story.append(ct)

        # Plan d'actions (si des actions existent)
        if acts:
            story.append(Spacer(1, 5*mm))
            story.append(HRFlowable(width='100%', thickness=1, color=BLUE))
            story.append(Paragraph("<b>PLAN D'ACTIONS</b>", h2))

            PRIO_COLORS = {'critical': '#FF0000', 'high': '#FF8C00', 'medium': '#1E4FA5', 'low': '#666666'}
            STATUT_FR   = {'todo': 'À faire', 'in_progress': 'En cours', 'done': 'Terminé'}
            PRIO_FR     = {'critical': 'Critique', 'high': 'Haute', 'medium': 'Moyenne', 'low': 'Faible'}

            ah = [Paragraph(f'<b>{h}</b>', sb) for h in
                  ['Critère','Titre','Responsable','Échéance','Statut','Priorité','Description']]
            adata = [ah]
            for act in acts:
                pc = PRIO_COLORS.get(act['priorite'], '#333333')
                adata.append([
                    Paragraph(act['critere_num'] or '', tiny),
                    Paragraph((act['titre'] or '')[:60], tiny),
                    Paragraph(act['responsable'] or '', tiny),
                    Paragraph(act['echeance'] or '', tiny),
                    Paragraph(STATUT_FR.get(act['statut'], act['statut']), tiny),
                    Paragraph(f'<font color="{pc}"><b>{PRIO_FR.get(act["priorite"], act["priorite"])}</b></font>', tiny),
                    Paragraph((act['description'] or '')[:100], tiny),
                ])
            act_table = Table(adata, colWidths=[2*cm, 5*cm, 3*cm, 2.5*cm, 2.5*cm, 2.5*cm, 8.5*cm], repeatRows=1)
            act_table.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),BLUE),
                ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTSIZE',(0,0),(-1,-1),6),
                ('VALIGN',(0,0),(-1,-1),'TOP'),
                ('INNERGRID',(0,0),(-1,-1),0.3,colors.lightgrey),
                ('BOX',(0,0),(-1,-1),0.5,colors.grey),
            ]))
            story.append(act_table)

        doc.build(story)
        output.seek(0)

        nom      = (e['nom'] or 'eval').replace(' ','_')
        filename = f"GoCertif_{nom}_{datetime.now():%Y%m%d}.pdf"
        self.set_header('Content-Type', 'application/pdf')
        self.set_header('Content-Disposition', f'attachment; filename="{filename}"')
        self.write(output.read())

# ─── Réponses EE ──────────────────────────────────────────────────────────────
class ReponsesEEHandler(BaseHandler):
    def get(self, eval_id):
        user = self.require_auth()
        if not user: return
        conn = get_db()
        e = conn.execute('SELECT id FROM evaluations WHERE id=? AND user_id=?',
                         (eval_id, user['id'])).fetchone()
        if not e:
            conn.close()
            return self.json({'error': 'Non trouvé'}, 404)
        rows = conn.execute(
            'SELECT critere_numero, ee_index, reponse FROM reponses_ee WHERE evaluation_id=?',
            (eval_id,)
        ).fetchall()
        conn.close()
        self.json([dict(r) for r in rows])

    def post(self, eval_id):
        """Bulk upsert : [{critere_numero, ee_index, reponse}, ...]"""
        user = self.require_auth()
        if not user: return
        conn = get_db()
        e = conn.execute('SELECT id FROM evaluations WHERE id=? AND user_id=?',
                         (eval_id, user['id'])).fetchone()
        if not e:
            conn.close()
            return self.json({'error': 'Non trouvé'}, 404)
        items = self.body_json()
        if not isinstance(items, list):
            conn.close()
            return self.json({'error': 'Liste attendue'}, 400)
        for item in items:
            num = item.get('critere_numero', '')
            idx = item.get('ee_index', 0)
            rep = item.get('reponse', '')
            if rep not in ('Oui', 'Non', 'NA', ''):
                continue
            conn.execute('''
                INSERT INTO reponses_ee (evaluation_id, critere_numero, ee_index, reponse, updated_at)
                VALUES (?,?,?,?,datetime('now'))
                ON CONFLICT(evaluation_id, critere_numero, ee_index)
                DO UPDATE SET reponse=excluded.reponse, updated_at=excluded.updated_at
            ''', (eval_id, num, idx, rep))
        conn.commit()
        conn.close()
        self.json({'ok': True})

# ─── IA — Suggestions plan d'actions ─────────────────────────────────────────
class AISuggestHandler(BaseHandler):
    def post(self):
        user = self.require_auth()
        if not user: return

        api_key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
        if not api_key:
            return self.json({'error': 'Clé API Anthropic non configurée (ANTHROPIC_API_KEY)'}, 503)

        data        = self.body_json()
        numero      = data.get('numero', '')
        critere     = data.get('critere', '')
        chapitre    = data.get('chapitre', '')
        objectif    = data.get('objectif', '')
        niveau      = data.get('niveau', '')
        champ       = data.get('champ', '')

        prompt = f"""Tu es un expert qualité hospitalière spécialisé dans la certification HAS V2025.

Un établissement de santé a coté "Non conforme" le critère suivant :
- Numéro : {numero}
- Chapitre : {chapitre}
- Objectif : {objectif}
- Critère : {critere}
- Niveau : {niveau}
- Champ : {champ}

Propose exactement 3 actions correctives concrètes, réalistes et directement applicables.
Chaque action doit être en 1 à 2 phrases maximum, avec un vocabulaire professionnel HAS.
Réponds uniquement avec une liste numérotée (1. 2. 3.), sans introduction ni conclusion."""

        import urllib.request as urlreq
        payload = json.dumps({
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 400,
            "messages": [{"role": "user", "content": prompt}]
        }).encode('utf-8')

        req = urlreq.Request(
            'https://api.anthropic.com/v1/messages',
            data=payload,
            headers={
                'x-api-key':           api_key,
                'anthropic-version':   '2023-06-01',
                'content-type':        'application/json',
            },
            method='POST'
        )
        try:
            with urlreq.urlopen(req, timeout=20) as resp:
                result = json.loads(resp.read().decode('utf-8'))
                text   = result['content'][0]['text'].strip()
                return self.json({'suggestions': text})
        except Exception as e:
            return self.json({'error': f'Erreur API : {e}'}, 500)

# ─── Application Tornado ───────────────────────────────────────────────────────
def make_app():
    return tornado.web.Application([
        (r'/',                                           IndexHandler),
        # Auth
        (r'/auth/register',                              AuthRegisterHandler),
        (r'/auth/login',                                 AuthLoginHandler),
        (r'/auth/logout',                                AuthLogoutHandler),
        (r'/auth/me',                                    AuthMeHandler),
        # Criteria
        (r'/api/criteria',                               CriteriaHandler),
        (r'/api/criteria/meta',                          CriteriaMetaHandler),
        # Evaluations
        (r'/api/evaluations',                            EvaluationsHandler),
        (r'/api/evaluations/import',                     ImportJSONHandler),
        (r'/api/evaluations/(\d+)',                      EvaluationHandler),
        (r'/api/evaluations/(\d+)/reponses',             ReponsesHandler),
        (r'/api/evaluations/(\d+)/reponses/bulk',        BulkReponsesHandler),
        (r'/api/evaluations/(\d+)/reponses/ee',          ReponsesEEHandler),
        (r'/api/evaluations/(\d+)/score',                ScoreHandler),
        (r'/api/evaluations/(\d+)/export/excel',         ExportExcelHandler),
        (r'/api/evaluations/(\d+)/export/pdf',           ExportPDFHandler),
        (r'/api/evaluations/(\d+)/export/json',          ExportJSONHandler),
        # Actions
        (r'/api/actions',                                ActionsHandler),
        (r'/api/actions/(\d+)',                          ActionHandler),
        # Score history
        (r'/api/score/history',                          ScoreHistoryHandler),
        # IA
        (r'/api/ai/suggest',                             AISuggestHandler),
    ],
    cookie_secret=COOKIE_SECRET,
    static_path=os.path.join(BASE_DIR, 'static'),
    xsrf_cookies=False)

if __name__ == '__main__':
    seed_admin()
    app = make_app()
    app.listen(PORT)
    print('=' * 60)
    print('  GoCertif Web v2 — Grilles d\'audit HAS V2025')
    print('=' * 60)
    print(f'  ➜  Ouvrez : http://localhost:{PORT}')
    print(f'  ➜  Base de données : {DB_PATH}')
    print(f'  ➜  Critères chargés : {len(CRITERIA)} (V2025)')
    print('  Ctrl+C pour arrêter')
    print('=' * 60)
    tornado.ioloop.IOLoop.current().start()
